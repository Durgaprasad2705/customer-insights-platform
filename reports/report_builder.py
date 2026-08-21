"""
PDF & Excel Report Builders — comprehensive report generation with
all charts, KPIs, tables, and ML insights.
"""

import datetime
import logging
from io import BytesIO

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import (
    HRFlowable, Image, PageBreak, Paragraph,
    SimpleDocTemplate, Spacer, Table, TableStyle,
)
from reportlab.lib import colors

from config import APP_NAME, COMPANY_NAME
from reports.report_templates import (
    C_INDIGO, C_VIOLET, C_GREEN, C_AMBER, C_DARK, C_NAVY,
    C_TEXT, C_MUTED, C_BG, C_BORDER, C_WHITE,
    get_table_style, get_pdf_styles, hex_color,
)
from reports.report_cache import get_cached_dataframe_agg

LOGGER = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────────────────
# Helper: convert chart PNG bytes to a PDF Image element
# ──────────────────────────────────────────────────────────────────────────────

def _chart_elem(png_bytes: bytes | None, width: float, height: float) -> list:
    """Return a list containing an Image flowable (or empty list on failure)."""
    if not png_bytes:
        return []
    try:
        img = Image(BytesIO(png_bytes), width=width, height=height)
        return [img, Spacer(1, 8)]
    except Exception as exc:
        LOGGER.warning("Failed to create Image flowable: %s", exc)
        return []


def _safe_image_or_placeholder(
    png_bytes: bytes | None,
    width: float,
    height: float,
    chart_label: str,
    styles: dict,
) -> object:
    """Return an Image flowable from PNG bytes, or a styled placeholder.

    This replaces all bare ``Paragraph("Chart unavailable")`` fallbacks
    with either the actual image or — if png_bytes is itself a fallback
    placeholder (always non-None thanks to chart_exporter) — the image
    from the fallback PNG. Only if the bytes are truly None or corrupt
    do we return a styled text paragraph.
    """
    if png_bytes:
        try:
            return Image(BytesIO(png_bytes), width=width, height=height)
        except Exception as exc:
            LOGGER.warning(
                "Image flowable creation failed for '%s': %s", chart_label, exc
            )

    # Absolute last-resort text fallback (should rarely trigger since
    # chart_exporter now generates placeholder PNGs)
    return Paragraph(
        f"<i><font color='#94A3B8' size='8'>"
        f"📊 {chart_label} — data not available for this dataset"
        f"</font></i>",
        styles["body"],
    )


# ══════════════════════════════════════════════════════════════════════════════
# PDF Report Builder
# ══════════════════════════════════════════════════════════════════════════════

class PDFReportBuilder:
    def __init__(
        self,
        df: pd.DataFrame,
        insights: list[dict],
        charts: dict,
        ml_data: dict | None = None,
    ):
        self.df = df
        self.insights = insights
        self.charts = charts
        self.ml_data = ml_data or {}
        self.buffer = BytesIO()
        self.styles = get_pdf_styles()
        self.PAGE_W = letter[0] - 80

        self.amt_col = next((c for c in ["TotalAmount", "Revenue", "Amount"] if c in df.columns), None)
        self.cust_col = next((c for c in ["CustomerID"] if c in df.columns), None)
        self.cat_col = next((c for c in ["Category"] if c in df.columns), None)
        self.brand_col = next((c for c in ["Brand"] if c in df.columns), None)
        self.date_col = next((c for c in ["PurchaseDate", "Date", "OrderDate"] if c in df.columns), None)
        self.region_col = next((c for c in ["Region", "Location", "State"] if c in df.columns), None)
        self.prod_col = next((c for c in ["ProductName", "Product"] if c in df.columns), None)
        self.rating_col = next((c for c in ["CustomerRating", "Rating"] if c in df.columns), None)
        self.profit_col = next((c for c in ["ProfitMargin", "Profit"] if c in df.columns), None)
        self.gender_col = next((c for c in ["Gender", "Sex"] if c in df.columns), None)
        self.qty_col = next((c for c in ["Quantity", "Units"] if c in df.columns), None)

        self.total_rev = float(df[self.amt_col].sum()) if self.amt_col else 0.0
        self.total_orders = len(df)
        self.avg_order = float(df[self.amt_col].mean()) if self.amt_col else 0.0
        self.unique_custs = df[self.cust_col].nunique() if self.cust_col else 0
        self.total_profit = float(df[self.profit_col].sum()) if self.profit_col else 0.0
        self.today = datetime.date.today()

    def build(self) -> bytes:
        doc = SimpleDocTemplate(
            self.buffer, pagesize=letter,
            rightMargin=40, leftMargin=40,
            topMargin=44, bottomMargin=44,
        )
        elems = []
        elems.extend(self._build_cover())
        elems.extend(self._build_kpis())
        elems.extend(self._build_insights())
        elems.extend(self._build_charts())
        elems.extend(self._build_monthly_breakdown())
        elems.extend(self._build_category_summary())
        elems.extend(self._build_top_customers())
        elems.extend(self._build_brand_summary())
        elems.extend(self._build_ml_insights())
        elems.extend(self._build_inventory_summary())
        elems.extend(self._build_transaction_preview())
        elems.extend(self._build_footer())

        doc.build(elems)
        self.buffer.seek(0)
        return self.buffer.getvalue()

    # ── Cover Page ────────────────────────────────────────────────────────

    def _build_cover(self) -> list:
        elems = [Spacer(1, 0.6 * inch)]
        banner_data = [[Paragraph(
            f"<font color='#6366F1'><b>{APP_NAME}</b></font>",
            self.styles["banner"],
        )]]
        banner = Table(banner_data, colWidths=[self.PAGE_W])
        banner.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, -1), colors.HexColor("#F0F1FF")),
            ("TOPPADDING",  (0, 0), (-1, -1), 18),
            ("BOTTOMPADDING",(0,0), (-1, -1), 18),
            ("LEFTPADDING", (0, 0), (-1, -1), 20),
            ("ROUNDEDCORNERS", [6]),
        ]))
        elems.append(banner)
        elems.append(Spacer(1, 14))

        elems.append(Paragraph("Executive Intelligence Report", self.styles["title"]))
        elems.append(Paragraph(
            f"Prepared for <b>{COMPANY_NAME}</b> &nbsp;|&nbsp; "
            f"{self.today.strftime('%B %d, %Y')}",
            self.styles["cov_sub"],
        ))
        elems.append(HRFlowable(width="100%", thickness=2, color=C_INDIGO, spaceAfter=18))

        def _kpi_cell(label: str, value: str, colour=C_INDIGO) -> Table:
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            st = getSampleStyleSheet()
            inner = Table(
                [[Paragraph(value, ParagraphStyle("KV", parent=st["Normal"], fontName="Helvetica-Bold", fontSize=18, textColor=colour, leading=22)),
                  Paragraph(label, ParagraphStyle("KL", parent=st["Normal"], fontName="Helvetica", fontSize=9, textColor=C_MUTED))]],
                colWidths=[None, None],
            )
            cell = Table([[inner]], colWidths=[(self.PAGE_W - 12) / 2])
            cell.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, -1), C_BG),
                ("BOX",           (0, 0), (-1, -1), 0.8, C_BORDER),
                ("TOPPADDING",    (0, 0), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ("LEFTPADDING",   (0, 0), (-1, -1), 14),
            ]))
            return cell

        kpi_row1 = Table([
            [_kpi_cell("Total Revenue",     f"${self.total_rev:,.0f}",   C_INDIGO),
             _kpi_cell("Total Orders",      f"{self.total_orders:,}",     C_GREEN)],
        ], colWidths=[(self.PAGE_W - 12) / 2, (self.PAGE_W - 12) / 2], hAlign="LEFT")
        kpi_row1.setStyle(TableStyle([("LEFTPADDING", (0,0),(-1,-1), 0), ("RIGHTPADDING",(0,0),(-1,-1), 0)]))
        elems.append(kpi_row1)
        elems.append(Spacer(1, 8))

        kpi_row2 = Table([
            [_kpi_cell("Avg Order Value",   f"${self.avg_order:,.2f}",   C_VIOLET),
             _kpi_cell("Unique Customers",  f"{self.unique_custs:,}",     C_AMBER)],
        ], colWidths=[(self.PAGE_W - 12) / 2, (self.PAGE_W - 12) / 2], hAlign="LEFT")
        kpi_row2.setStyle(TableStyle([("LEFTPADDING", (0,0),(-1,-1), 0), ("RIGHTPADDING",(0,0),(-1,-1), 0)]))
        elems.append(kpi_row2)
        elems.append(Spacer(1, 18))

        meta_rows = [["Dataset Rows", "Columns", "Date Range", "Report Generated"]]
        date_range = "–"
        if self.date_col:
            parsed = pd.to_datetime(self.df[self.date_col], errors="coerce")
            mn, mx = parsed.min(), parsed.max()
            if pd.notna(mn) and pd.notna(mx):
                date_range = f"{mn.strftime('%b %d, %Y')} → {mx.strftime('%b %d, %Y')}"
        meta_rows.append([
            f"{len(self.df):,}",
            f"{len(self.df.columns)}",
            date_range,
            self.today.strftime("%b %d, %Y %H:%M"),
        ])
        t_meta = Table(meta_rows, colWidths=[self.PAGE_W / 4] * 4)
        t_meta.setStyle(get_table_style(C_NAVY))
        elems.append(t_meta)
        elems.append(PageBreak())
        return elems

    # ── KPI Summary ───────────────────────────────────────────────────────

    def _build_kpis(self) -> list:
        elems = [Paragraph("1. Executive KPI Summary", self.styles["sec"])]
        kpi_data = [
            ["Metric", "Value", "Benchmark"],
            ["Total Revenue",         f"${self.total_rev:,.2f}",   "Active dataset"],
            ["Total Orders / Txns",   f"{self.total_orders:,}",    "Active dataset"],
            ["Average Order Value",   f"${self.avg_order:,.2f}",   "Calculated"],
            ["Unique Customers",      f"{self.unique_custs:,}",    "Active dataset"],
            ["Revenue per Customer",  f"${(self.total_rev / max(self.unique_custs,1)):,.2f}", "Calculated"],
            ["Orders per Customer",   f"{(self.total_orders / max(self.unique_custs,1)):.1f}", "Calculated"],
        ]

        # Additional KPIs when data is available
        if self.rating_col:
            avg_rating = float(self.df[self.rating_col].mean())
            kpi_data.append(["Avg Customer Rating", f"{avg_rating:.2f} / 5.0", "Calculated"])

        if self.profit_col:
            margin_pct = (self.total_profit / max(self.total_rev, 1)) * 100
            kpi_data.append(["Total Profit", f"${self.total_profit:,.2f}", "Active dataset"])
            kpi_data.append(["Profit Margin", f"{margin_pct:.1f}%", "Calculated"])

        if self.qty_col:
            total_units = int(self.df[self.qty_col].sum())
            kpi_data.append(["Total Units Sold", f"{total_units:,}", "Active dataset"])

        if self.cat_col:
            num_categories = self.df[self.cat_col].nunique()
            kpi_data.append(["Product Categories", f"{num_categories}", "Active dataset"])

        if self.brand_col:
            num_brands = self.df[self.brand_col].nunique()
            kpi_data.append(["Brands", f"{num_brands}", "Active dataset"])

        if self.region_col:
            num_regions = self.df[self.region_col].nunique()
            kpi_data.append(["Regions / Locations", f"{num_regions}", "Active dataset"])

        t_kpi = Table(kpi_data, colWidths=[230, 160, 150])
        t_kpi.setStyle(get_table_style(C_INDIGO))
        elems.extend([t_kpi, Spacer(1, 16)])
        return elems

    # ── AI Insights ───────────────────────────────────────────────────────

    def _build_insights(self) -> list:
        elems = [Paragraph("2. Strategic AI Insights", self.styles["sec"])]
        ins_data = [["Priority", "Insight", "Recommended Action", "Confidence"]]
        for item in self.insights[:5]:
            p = item.get("priority", "–")
            col = {"High": C_INDIGO, "Medium": C_AMBER, "Low": C_GREEN}.get(p, C_TEXT)
            ins_data.append([
                Paragraph(f"<font color='#{hex_color(col)}'><b>{p}</b></font>", self.styles["body"]),
                Paragraph(item.get("title",    "–"), self.styles["body"]),
                Paragraph(item.get("recommended_action", "–"), self.styles["body"]),
                f"{item.get('confidence', 0):.0f}%",
            ])
        t_ins = Table(ins_data, colWidths=[58, 160, 250, 72])
        t_ins.setStyle(get_table_style(C_DARK))
        elems.extend([t_ins, Spacer(1, 16)])
        return elems

    # ── Dashboard Charts ──────────────────────────────────────────────────

    def _build_charts(self) -> list:
        elems = [PageBreak(), Paragraph("3. Dashboard Charts", self.styles["sec"])]
        pngs = self.charts
        half = (self.PAGE_W - 8) / 2

        # 3a. Monthly Revenue Trend (full width)
        elems.append(Paragraph("3a. Monthly Revenue Trend", self.styles["sub_sec"]))
        png = pngs.get("trend")
        elems += _chart_elem(png, self.PAGE_W, 3.2 * inch)
        if png:
            elems.append(Paragraph(
                "Monthly revenue aggregated by calendar month from the active dataset.",
                self.styles["caption"],
            ))

        # 3b. Category Revenue & Brand Performance (side-by-side)
        elems.append(Paragraph("3b. Revenue by Category & Brand Performance", self.styles["sub_sec"]))
        row_data = [[
            _safe_image_or_placeholder(pngs.get("cat"), half, 2.7 * inch, "Category Revenue", self.styles),
            _safe_image_or_placeholder(pngs.get("brand"), half, 2.7 * inch, "Brand Performance", self.styles),
        ]]
        side_tbl = Table(row_data, colWidths=[half, half])
        side_tbl.setStyle(TableStyle([
            ("LEFTPADDING", (0,0),(-1,-1), 0), ("RIGHTPADDING", (0,0),(-1,-1), 0),
            ("TOPPADDING", (0,0),(-1,-1), 0), ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ]))
        elems.append(side_tbl)

        # 3c. Regional Sales & Top Products (side-by-side)
        elems.append(Paragraph("3c. Regional Sales & Top Products", self.styles["sub_sec"]))
        row_data2 = [[
            _safe_image_or_placeholder(pngs.get("reg"), half, 2.7 * inch, "Regional Sales", self.styles),
            _safe_image_or_placeholder(pngs.get("prod"), half, 2.7 * inch, "Top Products", self.styles),
        ]]
        side_tbl2 = Table(row_data2, colWidths=[half, half])
        side_tbl2.setStyle(TableStyle([
            ("LEFTPADDING", (0,0),(-1,-1), 0), ("RIGHTPADDING", (0,0),(-1,-1), 0),
            ("TOPPADDING", (0,0),(-1,-1), 0), ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ]))
        elems.append(side_tbl2)

        # 3d. Payment Methods & Customer Age Distribution (side-by-side)
        elems.append(PageBreak())
        elems.append(Paragraph("3d. Payment Methods & Customer Age Distribution", self.styles["sub_sec"]))
        row_data3 = [[
            _safe_image_or_placeholder(pngs.get("pay"), half, 2.7 * inch, "Payment Methods", self.styles),
            _safe_image_or_placeholder(pngs.get("age"), half, 2.7 * inch, "Age Distribution", self.styles),
        ]]
        side_tbl3 = Table(row_data3, colWidths=[half, half])
        side_tbl3.setStyle(TableStyle([
            ("LEFTPADDING", (0,0),(-1,-1), 0), ("RIGHTPADDING", (0,0),(-1,-1), 0),
            ("TOPPADDING", (0,0),(-1,-1), 0), ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ]))
        elems.append(side_tbl3)

        # 3e. Customer Conversion Funnel (wide)
        elems.append(Paragraph("3e. Customer Conversion Funnel", self.styles["sub_sec"]))
        png_funnel = pngs.get("funnel")
        elems += _chart_elem(png_funnel, self.PAGE_W * 0.75, 2.8 * inch)

        # 3f. Gender Distribution (if available)
        if "gender" in pngs:
            elems.append(Paragraph("3f. Customer Gender Distribution", self.styles["sub_sec"]))
            elems += _chart_elem(pngs.get("gender"), self.PAGE_W * 0.55, 2.7 * inch)

        # 3g. Customer Segments & Revenue Forecast (ML charts, side-by-side)
        has_seg = "segment" in pngs
        has_fore = "forecast" in pngs
        if has_seg or has_fore:
            elems.append(PageBreak())
            section_label = "3g" if "gender" in pngs else "3f"
            elems.append(Paragraph(
                f"{section_label}. Customer Segments & Revenue Forecast",
                self.styles["sub_sec"],
            ))
            row_data4 = [[
                _safe_image_or_placeholder(
                    pngs.get("segment"), half, 2.7 * inch,
                    "Customer Segments", self.styles,
                ),
                _safe_image_or_placeholder(
                    pngs.get("forecast"), half, 2.7 * inch,
                    "Revenue Forecast", self.styles,
                ),
            ]]
            side_tbl4 = Table(row_data4, colWidths=[half, half])
            side_tbl4.setStyle(TableStyle([
                ("LEFTPADDING", (0,0),(-1,-1), 0), ("RIGHTPADDING", (0,0),(-1,-1), 0),
                ("TOPPADDING", (0,0),(-1,-1), 0), ("BOTTOMPADDING",(0,0),(-1,-1), 4),
            ]))
            elems.append(side_tbl4)

        return elems

    # ── Monthly Revenue Breakdown ─────────────────────────────────────────

    def _build_monthly_breakdown(self) -> list:
        elems = [PageBreak(), Paragraph("4. Monthly Revenue Breakdown", self.styles["sec"])]
        if self.date_col and self.amt_col:
            def _agg(df):
                tmp = df[[self.date_col, self.amt_col]].copy()
                tmp[self.date_col] = pd.to_datetime(tmp[self.date_col], errors="coerce")
                tmp[self.amt_col]  = pd.to_numeric(tmp[self.amt_col], errors="coerce")
                tmp = tmp.dropna()
                if tmp.empty: return None
                return tmp.groupby(pd.Grouper(key=self.date_col, freq="ME"))[self.amt_col].agg(Revenue="sum", Orders="count").reset_index()

            monthly = get_cached_dataframe_agg(self.df, "monthly_rev", _agg)

            if monthly is not None and not monthly.empty:
                m2 = monthly.copy()
                m2.columns = ["Month", "Revenue", "Orders"]
                m2["Avg Order"] = (m2["Revenue"] / m2["Orders"].replace(0, 1)).round(2)
                m2 = m2.sort_values("Month", ascending=False)
                m2["Month"]   = m2["Month"].dt.strftime("%b %Y")
                m2["Revenue"] = m2["Revenue"].apply(lambda v: f"${v:,.2f}")
                m2["Avg Order"] = m2["Avg Order"].apply(lambda v: f"${v:,.2f}")
                mon_data = [list(m2.columns)] + m2.values.tolist()
                t_mon = Table(mon_data, colWidths=[120, 160, 100, 160])
                t_mon.setStyle(get_table_style(C_GREEN))
                elems.extend([t_mon, Spacer(1, 16)])
            else:
                elems.append(Paragraph("Date/Amount columns not found for monthly breakdown.", self.styles["body"]))
        return elems

    # ── Category Summary ──────────────────────────────────────────────────

    def _build_category_summary(self) -> list:
        elems = []
        if self.cat_col and self.amt_col:
            elems.append(Paragraph("5. Revenue by Category", self.styles["sec"]))
            def _agg(df): return df.groupby(self.cat_col)[self.amt_col].agg(["sum", "count", "mean"]).reset_index()
            cat_df = get_cached_dataframe_agg(self.df, "cat_summary", _agg).copy()
            cat_df.columns = ["Category", "Total Revenue", "Orders", "Avg Order Value"]
            cat_df = cat_df.sort_values("Total Revenue", ascending=False)
            cat_df["Total Revenue"]    = cat_df["Total Revenue"].apply(lambda v: f"${v:,.2f}")
            cat_df["Avg Order Value"]  = cat_df["Avg Order Value"].apply(lambda v: f"${v:,.2f}")
            cat_data = [list(cat_df.columns)] + cat_df.values.tolist()
            t_cat = Table(cat_data, colWidths=[160, 140, 100, 140])
            t_cat.setStyle(get_table_style(C_VIOLET))
            elems.extend([t_cat, Spacer(1, 16)])
        return elems

    # ── Top Customers ─────────────────────────────────────────────────────

    def _build_top_customers(self) -> list:
        elems = []
        if self.cust_col and self.amt_col:
            elems.append(Paragraph("6. Top 10 Customers by Revenue", self.styles["sec"]))
            def _agg(df): return df.groupby(self.cust_col)[self.amt_col].sum().sort_values(ascending=False).head(10).reset_index()
            top_custs = get_cached_dataframe_agg(self.df, "top_cust_10", _agg).copy()
            top_custs.columns = ["Customer ID", "Total Revenue"]
            top_custs.insert(0, "Rank", range(1, len(top_custs) + 1))
            top_custs["Total Revenue"] = top_custs["Total Revenue"].apply(lambda v: f"${v:,.2f}")
            tc_data = [list(top_custs.columns)] + top_custs.values.tolist()
            t_tc = Table(tc_data, colWidths=[60, 220, 180])
            t_tc.setStyle(get_table_style(C_AMBER))
            elems.extend([t_tc, Spacer(1, 16)])
        return elems

    # ── Brand Summary ─────────────────────────────────────────────────────

    def _build_brand_summary(self) -> list:
        elems = []
        if self.brand_col and self.amt_col:
            elems.append(Paragraph("7. Brand Performance Summary", self.styles["sec"]))
            def _agg(df): return df.groupby(self.brand_col)[self.amt_col].agg(["sum", "count"]).reset_index()
            brand_df = get_cached_dataframe_agg(self.df, "brand_summary", _agg).copy()
            brand_df.columns = ["Brand", "Total Revenue", "Orders"]
            brand_df = brand_df.sort_values("Total Revenue", ascending=False).head(15)
            brand_df["Total Revenue"] = brand_df["Total Revenue"].apply(lambda v: f"${v:,.2f}")
            br_data = [list(brand_df.columns)] + brand_df.values.tolist()
            t_br = Table(br_data, colWidths=[200, 160, 100])
            t_br.setStyle(get_table_style(colors.HexColor("#0EA5E9")))
            elems.extend([t_br, Spacer(1, 16)])
        return elems

    # ── ML Insights Section ───────────────────────────────────────────────

    def _build_ml_insights(self) -> list:
        """Add ML model outputs: segment summary and forecast table."""
        elems = []
        segment_summary = self.ml_data.get("segment_summary")
        forecast_data = self.ml_data.get("forecast_data")

        if segment_summary is None and forecast_data is None:
            return elems

        elems.append(PageBreak())
        elems.append(Paragraph("8. Machine Learning Insights", self.styles["sec"]))

        # Segment profiles table
        if segment_summary is not None and not segment_summary.empty:
            elems.append(Paragraph("8a. Customer Segment Profiles", self.styles["sub_sec"]))
            seg_display = segment_summary.copy()
            # Format numeric columns
            for col in ["AvgRecency", "AvgFrequency", "AvgMonetary", "TotalRevenue"]:
                if col in seg_display.columns:
                    if col in ("AvgMonetary", "TotalRevenue"):
                        seg_display[col] = seg_display[col].apply(lambda v: f"${v:,.0f}")
                    else:
                        seg_display[col] = seg_display[col].apply(lambda v: f"{v:.1f}")
            # Select display columns
            show_cols = [c for c in ["SegmentName", "CustomerCount", "AvgRecency",
                                      "AvgFrequency", "AvgMonetary", "TotalRevenue"]
                         if c in seg_display.columns]
            if show_cols:
                seg_show = seg_display[show_cols]
                seg_data = [list(seg_show.columns)] + seg_show.values.tolist()
                col_widths = [int(self.PAGE_W / len(show_cols))] * len(show_cols)
                t_seg = Table(seg_data, colWidths=col_widths)
                t_seg.setStyle(get_table_style(C_INDIGO))
                elems.extend([t_seg, Spacer(1, 16)])

        # Revenue forecast table
        if forecast_data is not None and not forecast_data.empty:
            elems.append(Paragraph("8b. Revenue Forecast (Projected)", self.styles["sub_sec"]))
            fore_display = forecast_data.copy()
            if "Month" in fore_display.columns:
                fore_display["Month"] = pd.to_datetime(fore_display["Month"], errors="coerce").dt.strftime("%b %Y")
            if "Revenue" in fore_display.columns:
                fore_display["Revenue"] = fore_display["Revenue"].apply(lambda v: f"${v:,.0f}")
            show_cols = [c for c in ["Month", "Revenue"] if c in fore_display.columns]
            if show_cols:
                fore_show = fore_display[show_cols]
                fore_data_rows = [list(fore_show.columns)] + fore_show.values.tolist()
                t_fore = Table(fore_data_rows, colWidths=[200, 200])
                t_fore.setStyle(get_table_style(C_GREEN))
                elems.extend([t_fore, Spacer(1, 16)])

        return elems

    # ── Inventory Summary ─────────────────────────────────────────────────

    def _build_inventory_summary(self) -> list:
        """Add product velocity and stock-level summary from the dataset."""
        elems = []
        if not self.prod_col or not self.qty_col:
            return elems

        elems.append(Paragraph("9. Product Inventory & Velocity Summary", self.styles["sec"]))

        grp_cols = [c for c in [self.cat_col, self.prod_col] if c]
        agg_dict = {"UnitsSold": (self.qty_col, "sum")}
        if self.amt_col:
            agg_dict["Revenue"] = (self.amt_col, "sum")

        def _agg(df):
            inv = df.groupby(grp_cols).agg(**{k: (v[0], v[1]) for k, v in agg_dict.items()}).reset_index()
            inv = inv.sort_values("UnitsSold", ascending=False).head(15)
            return inv

        inv_df = get_cached_dataframe_agg(self.df, "inventory_summary_pdf", _agg)
        if inv_df is not None and not inv_df.empty:
            display_df = inv_df.copy()
            if "Revenue" in display_df.columns:
                display_df["Revenue"] = display_df["Revenue"].apply(lambda v: f"${v:,.0f}")
            inv_data = [list(display_df.columns)] + display_df.values.tolist()
            col_widths = [int(self.PAGE_W / len(display_df.columns))] * len(display_df.columns)
            t_inv = Table(inv_data, colWidths=col_widths)
            t_inv.setStyle(get_table_style(colors.HexColor("#8B5CF6")))
            elems.extend([t_inv, Spacer(1, 16)])

        return elems

    # ── Transaction Preview ───────────────────────────────────────────────

    def _build_transaction_preview(self) -> list:
        elems = [PageBreak(), Paragraph("10. Transaction Data Preview (Top 20 rows)", self.styles["sec"])]
        show_cols = [c for c in ["TransactionID", "CustomerID", "PurchaseDate", "Category", "ProductName", "TotalAmount", "PaymentMethod"] if c in self.df.columns]
        sample   = self.df[show_cols].head(20).copy()
        if self.date_col and self.date_col in sample.columns:
            sample[self.date_col] = pd.to_datetime(sample[self.date_col], errors="coerce").dt.strftime("%Y-%m-%d")
        if self.amt_col and self.amt_col in sample.columns:
            sample[self.amt_col] = sample[self.amt_col].apply(lambda v: f"${float(v):,.2f}")
        col_w = [int(self.PAGE_W / max(1, len(show_cols)))] * len(show_cols)
        tbl_data = [list(sample.columns)] + [[str(v)[:22] for v in row] for _, row in sample.iterrows()]
        t_txn = Table(tbl_data, colWidths=col_w)
        t_txn.setStyle(get_table_style(colors.HexColor("#334155")))
        elems.append(t_txn)
        return elems

    # ── Footer ────────────────────────────────────────────────────────────

    def _build_footer(self) -> list:
        return [
            Spacer(1, 20),
            HRFlowable(width="100%", thickness=0.8, color=C_BORDER, spaceAfter=8),
            Paragraph(
                f"<i>This report was auto-generated by {APP_NAME} on "
                f"{self.today.strftime('%B %d, %Y')}. "
                f"Data reflects the currently active filtered dataset ({len(self.df):,} rows).</i>",
                self.styles["footer"]
            )
        ]


# ══════════════════════════════════════════════════════════════════════════════
# Excel Report Builder
# ══════════════════════════════════════════════════════════════════════════════

class ExcelReportBuilder:
    def __init__(self, df: pd.DataFrame, insights: list[dict]):
        self.df = df
        self.insights = insights
        self.wb = openpyxl.Workbook()
        self.amt_col = next((c for c in ["TotalAmount", "Revenue", "Amount"] if c in df.columns), None)

        self.fill_hdr    = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        self.fill_hdr2   = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        self.fill_hdr3   = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        self.fill_hdr4   = PatternFill(start_color="A855F7", end_color="A855F7", fill_type="solid")
        self.fill_hdr5   = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        self.fill_hdr6   = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
        self.fill_alt    = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        self.font_hdr    = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        self.font_body   = Font(name="Segoe UI", size=10, color="1E293B")
        self.font_title  = Font(name="Segoe UI", size=14, bold=True, color="6366F1")
        self.thin = Border(
            left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0")
        )
        self.center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def _auto_width(self, ws) -> None:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = min(max_len + 4, 42)

    def _write_sheet(self, ws, frame: pd.DataFrame, hdr_fill=None, title: str = "") -> None:
        if hdr_fill is None: hdr_fill = self.fill_hdr
        if title:
            ws.append([title])
            title_cell = ws.cell(row=1, column=1)
            title_cell.font = self.font_title
            title_cell.alignment = self.left
            ws.append([])
            start_row = 3
        else:
            start_row = 1
        for r_idx, row in enumerate(dataframe_to_rows(frame, index=False, header=True), start_row):
            ws.append(row)
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(row=r_idx if title else r_idx, column=c_idx)
                if r_idx == start_row:
                    cell.fill, cell.font, cell.alignment = hdr_fill, self.font_hdr, self.center
                else:
                    cell.font, cell.border, cell.alignment = self.font_body, self.thin, self.left
                    if (r_idx - start_row) % 2 == 1: cell.fill = self.fill_alt
        self._auto_width(ws)

    def build(self) -> bytes:
        ws1 = self.wb.active
        ws1.title = "Transaction Data"
        ws1.sheet_view.showGridLines = False
        self._write_sheet(ws1, self.df, self.fill_hdr, title=f"{APP_NAME} – Transaction Data")

        date_col = next((c for c in ["PurchaseDate", "Date", "OrderDate"] if c in self.df.columns), None)
        if date_col and self.amt_col:
            def _agg_m(df):
                tmp = df[[date_col, self.amt_col]].copy()
                tmp[date_col] = pd.to_datetime(tmp[date_col], errors="coerce")
                tmp[self.amt_col] = pd.to_numeric(tmp[self.amt_col], errors="coerce")
                tmp = tmp.dropna()
                if tmp.empty: return None
                return tmp.groupby(pd.Grouper(key=date_col, freq="ME"))[self.amt_col].agg(Revenue="sum", Orders="count").reset_index()

            monthly = get_cached_dataframe_agg(self.df, "monthly_rev", _agg_m)
            if monthly is not None and not monthly.empty:
                m2 = monthly.copy()
                m2.columns = ["Month", "Revenue", "Orders"]
                m2["Avg Order Value"] = (m2["Revenue"] / m2["Orders"].replace(0, 1)).round(2)
                m2["Month"] = m2["Month"].dt.strftime("%b %Y")
                m2 = m2.sort_values("Month")
                ws2 = self.wb.create_sheet("Monthly Revenue")
                ws2.sheet_view.showGridLines = False
                self._write_sheet(ws2, m2, self.fill_hdr3, title="Monthly Revenue Breakdown")

        cat_col = next((c for c in ["Category"] if c in self.df.columns), None)
        if cat_col and self.amt_col:
            def _agg_c(df): return df.groupby(cat_col).agg(Orders=(self.amt_col, "count"), Revenue=(self.amt_col, "sum"), AvgBasket=(self.amt_col, "mean")).reset_index()
            cat = get_cached_dataframe_agg(self.df, "cat_summary_excel", _agg_c).copy()
            cat.columns = ["Category", "Orders", "Revenue", "Avg Basket"]
            cat = cat.sort_values("Revenue", ascending=False)
            ws3 = self.wb.create_sheet("Category Summary")
            ws3.sheet_view.showGridLines = False
            self._write_sheet(ws3, cat, self.fill_hdr4, title="Revenue by Category")

        brand_col = next((c for c in ["Brand"] if c in self.df.columns), None)
        if brand_col and self.amt_col:
            def _agg_b(df): return df.groupby(brand_col).agg(Orders=(self.amt_col, "count"), Revenue=(self.amt_col, "sum"), AvgBasket=(self.amt_col, "mean")).reset_index()
            brand_df = get_cached_dataframe_agg(self.df, "brand_summary_excel", _agg_b).copy()
            brand_df.columns = ["Brand", "Orders", "Revenue", "Avg Basket"]
            brand_df = brand_df.sort_values("Revenue", ascending=False)
            ws4 = self.wb.create_sheet("Brand Summary")
            ws4.sheet_view.showGridLines = False
            self._write_sheet(ws4, brand_df, self.fill_hdr6, title="Brand Performance")

        region_col = next((c for c in ["Region", "Location", "State"] if c in self.df.columns), None)
        if region_col and self.amt_col:
            def _agg_r(df): return df.groupby(region_col).agg(Orders=(self.amt_col, "count"), Revenue=(self.amt_col, "sum")).reset_index()
            reg_df = get_cached_dataframe_agg(self.df, "reg_summary_excel", _agg_r).copy()
            reg_df.columns = ["Region", "Orders", "Revenue"]
            reg_df = reg_df.sort_values("Revenue", ascending=False)
            ws5 = self.wb.create_sheet("Regional Summary")
            ws5.sheet_view.showGridLines = False
            self._write_sheet(ws5, reg_df, self.fill_hdr2, title="Sales by Region")

        cust_col = next((c for c in ["CustomerID"] if c in self.df.columns), None)
        if cust_col and self.amt_col:
            rating_col = next((c for c in ["CustomerRating", "Rating"] if c in self.df.columns), None)
            def _agg_cu(df):
                agg_dict = {"TotalSpend": (self.amt_col, "sum"), "OrderCount": (self.amt_col, "count"), "AvgOrder": (self.amt_col, "mean")}
                if rating_col: agg_dict["AvgRating"] = (rating_col, "mean")
                return df.groupby(cust_col).agg(**agg_dict).reset_index()

            cust_summary = get_cached_dataframe_agg(self.df, "cust_summary_excel", _agg_cu).copy()
            cust_summary = cust_summary.sort_values("TotalSpend", ascending=False).head(50)
            ws6 = self.wb.create_sheet("Customer Summary")
            ws6.sheet_view.showGridLines = False
            self._write_sheet(ws6, cust_summary, self.fill_hdr5, title="Top 50 Customers by Spend")

        if self.insights:
            ws7 = self.wb.create_sheet("AI Insights")
            ws7.sheet_view.showGridLines = False
            ins_df = pd.DataFrame([{
                "Priority":    i.get("priority",    ""),
                "Insight":     i.get("title",       ""),
                "Description": i.get("description", ""),
                "Action":      i.get("recommended_action", ""),
                "Impact":      i.get("impact",      ""),
                "Confidence":  f"{i.get('confidence', 0):.0f}%",
            } for i in self.insights])
            self._write_sheet(ws7, ins_df, self.fill_hdr2, title="AI Strategic Insights")

        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
