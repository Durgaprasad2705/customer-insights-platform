"""
Customer Insights Platform – Report Generation Service
Produces PDF (ReportLab + embedded Plotly charts), Excel (OpenPyXL), and CSV exports.
"""

from __future__ import annotations

import datetime
import logging
from io import BytesIO
from typing import Any

import pandas as pd

# ─── ReportLab ───────────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    HRFlowable, Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer,
    Table, TableStyle,
)

# ─── OpenPyXL ────────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from config import APP_NAME, COMPANY_NAME

LOGGER = logging.getLogger(__name__)

# ─── Colour palette ──────────────────────────────────────────────────────────
C_INDIGO  = colors.HexColor("#6366F1")
C_VIOLET  = colors.HexColor("#A855F7")
C_GREEN   = colors.HexColor("#10B981")
C_AMBER   = colors.HexColor("#F59E0B")
C_DARK    = colors.HexColor("#0F172A")
C_NAVY    = colors.HexColor("#1E293B")
C_TEXT    = colors.HexColor("#334155")
C_MUTED   = colors.HexColor("#64748B")
C_BG      = colors.HexColor("#F8FAFC")
C_BG2     = colors.HexColor("#F1F5F9")
C_BORDER  = colors.HexColor("#E2E8F0")
C_WHITE   = colors.white


# ─── Chart → PNG helper ───────────────────────────────────────────────────────

def _fig_to_png(fig, width_px: int = 700, height_px: int = 380) -> bytes | None:
    """Convert a Plotly figure to PNG bytes using kaleido. Returns None on failure."""
    try:
        import plotly.io as pio
        return pio.to_image(
            fig, format="png", width=width_px, height=height_px, scale=1.5
        )
    except Exception as exc:
        LOGGER.warning("Chart render failed: %s", exc)
        return None


def _chart_elem(png_bytes: bytes | None, width: float, height: float) -> list:
    """Return a list containing an Image flowable (or empty list on failure)."""
    if not png_bytes:
        return []
    img = Image(BytesIO(png_bytes), width=width, height=height)
    return [img, Spacer(1, 8)]


# ─── Shared Table Style ───────────────────────────────────────────────────────

def _ts(header_color: Any, alt: bool = True) -> TableStyle:
    cmds = [
        ("BACKGROUND",    (0, 0), (-1, 0),  header_color),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  C_WHITE),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("GRID",          (0, 0), (-1, -1), 0.4, C_BORDER),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 8.5),
        ("TEXTCOLOR",     (0, 1), (-1, -1), C_TEXT),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [C_BG, C_BG2] if alt else [C_BG]),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]
    return TableStyle(cmds)


# ─── PDF Report ───────────────────────────────────────────────────────────────

def generate_pdf_report(df: pd.DataFrame, insights: list[dict]) -> bytes:
    """
    Generate a professional multi-page executive PDF report with:
      • Cover page
      • Executive KPI summary
      • Strategic AI insights
      • Dashboard charts (revenue trend, category, brand, region, top products, funnel)
      • Monthly revenue breakdown table
      • Top-10 customers table
      • Transaction preview
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=40, leftMargin=40,
        topMargin=44, bottomMargin=44,
    )

    styles    = getSampleStyleSheet()
    PAGE_W    = letter[0] - 80   # usable width

    title_style = ParagraphStyle(
        "RPTitle", parent=styles["Heading1"],
        fontName="Helvetica-Bold", fontSize=26,
        textColor=C_INDIGO, leading=30, spaceAfter=6,
    )
    cov_sub = ParagraphStyle(
        "CovSub", parent=styles["Normal"],
        fontName="Helvetica", fontSize=12,
        textColor=C_MUTED, spaceAfter=6,
    )
    sec_style = ParagraphStyle(
        "RPSec", parent=styles["Heading2"],
        fontName="Helvetica-Bold", fontSize=13,
        textColor=C_INDIGO, spaceBefore=16, spaceAfter=8,
    )
    sub_sec = ParagraphStyle(
        "RPSubSec", parent=styles["Heading3"],
        fontName="Helvetica-Bold", fontSize=10,
        textColor=C_NAVY, spaceBefore=10, spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "RPBody", parent=styles["Normal"],
        fontName="Helvetica", fontSize=9,
        textColor=C_TEXT, leading=13,
    )
    caption_style = ParagraphStyle(
        "RPCaption", parent=styles["Normal"],
        fontName="Helvetica-Oblique", fontSize=8,
        textColor=C_MUTED, spaceAfter=6, alignment=1,  # centred
    )
    kpi_val_style = ParagraphStyle(
        "KPIVal", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=16,
        textColor=C_INDIGO, leading=20,
    )
    kpi_lbl_style = ParagraphStyle(
        "KPILbl", parent=styles["Normal"],
        fontName="Helvetica", fontSize=8,
        textColor=C_MUTED,
    )

    elems: list = []

    # ── Pre-compute KPIs ─────────────────────────────────────────────────────
    amt_col  = next((c for c in ["TotalAmount", "Revenue", "Amount"] if c in df.columns), None)
    cust_col = next((c for c in ["CustomerID"] if c in df.columns), None)
    cat_col  = next((c for c in ["Category"]   if c in df.columns), None)
    brand_col= next((c for c in ["Brand"]       if c in df.columns), None)
    date_col = next((c for c in ["PurchaseDate", "Date", "OrderDate"] if c in df.columns), None)
    region_col=next((c for c in ["Region", "Location", "State"] if c in df.columns), None)
    prod_col = next((c for c in ["ProductName", "Product"] if c in df.columns), None)
    pay_col  = next((c for c in ["PaymentMethod", "Payment"] if c in df.columns), None)
    rating_col=next((c for c in ["CustomerRating", "Rating"] if c in df.columns), None)

    total_rev    = float(df[amt_col].sum())     if amt_col  else 0.0
    total_orders = len(df)
    avg_order    = float(df[amt_col].mean())    if amt_col  else 0.0
    unique_custs = df[cust_col].nunique()       if cust_col else 0
    today        = datetime.date.today()

    # ── COVER PAGE ────────────────────────────────────────────────────────────
    elems.append(Spacer(1, 0.6 * inch))

    # Gradient-like banner using a coloured Table cell
    banner_data = [[Paragraph(
        f"<font color='#6366F1'><b>{APP_NAME}</b></font>",
        ParagraphStyle("Ban", parent=styles["Normal"],
                       fontName="Helvetica-Bold", fontSize=28,
                       textColor=C_INDIGO, leading=34),
    )]]
    banner = Table(banner_data, colWidths=[PAGE_W])
    banner.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, -1), colors.HexColor("#F0F1FF")),
        ("TOPPADDING",  (0, 0), (-1, -1), 18),
        ("BOTTOMPADDING",(0,0), (-1, -1), 18),
        ("LEFTPADDING", (0, 0), (-1, -1), 20),
        ("ROUNDEDCORNERS", [6]),
    ]))
    elems.append(banner)
    elems.append(Spacer(1, 14))

    elems.append(Paragraph("Executive Intelligence Report", title_style))
    elems.append(Paragraph(
        f"Prepared for <b>{COMPANY_NAME}</b> &nbsp;|&nbsp; "
        f"{today.strftime('%B %d, %Y')}",
        cov_sub,
    ))
    elems.append(HRFlowable(width="100%", thickness=2, color=C_INDIGO, spaceAfter=18))

    # Cover KPI summary cards (2×2 grid)
    def _kpi_cell(label: str, value: str, colour: Any = C_INDIGO) -> Table:
        inner = Table(
            [[Paragraph(value, ParagraphStyle("KV", parent=styles["Normal"],
                        fontName="Helvetica-Bold", fontSize=18,
                        textColor=colour, leading=22)),
              Paragraph(label, ParagraphStyle("KL", parent=styles["Normal"],
                        fontName="Helvetica", fontSize=9,
                        textColor=C_MUTED))]],
            colWidths=[None, None],
        )
        cell = Table([[inner]], colWidths=[(PAGE_W - 12) / 2])
        cell.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), C_BG),
            ("BOX",           (0, 0), (-1, -1), 0.8, C_BORDER),
            ("TOPPADDING",    (0, 0), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ("LEFTPADDING",   (0, 0), (-1, -1), 14),
        ]))
        return cell

    kpi_row1 = Table([
        [_kpi_cell("Total Revenue",     f"${total_rev:,.0f}",   C_INDIGO),
         _kpi_cell("Total Orders",      f"{total_orders:,}",     C_GREEN)],
    ], colWidths=[(PAGE_W - 12) / 2, (PAGE_W - 12) / 2], hAlign="LEFT")
    kpi_row1.setStyle(TableStyle([("LEFTPADDING", (0,0),(-1,-1), 0),
                                   ("RIGHTPADDING",(0,0),(-1,-1), 0)]))
    elems.append(kpi_row1)
    elems.append(Spacer(1, 8))

    kpi_row2 = Table([
        [_kpi_cell("Avg Order Value",   f"${avg_order:,.2f}",   C_VIOLET),
         _kpi_cell("Unique Customers",  f"{unique_custs:,}",     C_AMBER)],
    ], colWidths=[(PAGE_W - 12) / 2, (PAGE_W - 12) / 2], hAlign="LEFT")
    kpi_row2.setStyle(TableStyle([("LEFTPADDING", (0,0),(-1,-1), 0),
                                   ("RIGHTPADDING",(0,0),(-1,-1), 0)]))
    elems.append(kpi_row2)
    elems.append(Spacer(1, 18))

    # Dataset metadata row
    meta_rows = [["Dataset Rows", "Columns", "Date Range", "Report Generated"]]
    date_range = "–"
    if date_col:
        parsed = pd.to_datetime(df[date_col], errors="coerce")
        mn, mx = parsed.min(), parsed.max()
        if pd.notna(mn) and pd.notna(mx):
            date_range = f"{mn.strftime('%b %d, %Y')} → {mx.strftime('%b %d, %Y')}"
    meta_rows.append([
        f"{len(df):,}",
        f"{len(df.columns)}",
        date_range,
        today.strftime("%b %d, %Y %H:%M"),
    ])
    t_meta = Table(meta_rows, colWidths=[PAGE_W / 4] * 4)
    t_meta.setStyle(_ts(C_NAVY))
    elems.append(t_meta)

    elems.append(PageBreak())

    # ── SECTION 1: EXECUTIVE KPIS ────────────────────────────────────────────
    elems.append(Paragraph("1. Executive KPI Summary", sec_style))
    kpi_data = [
        ["Metric", "Value", "Benchmark"],
        ["Total Revenue",         f"${total_rev:,.2f}",   "Active dataset"],
        ["Total Orders / Txns",   f"{total_orders:,}",    "Active dataset"],
        ["Average Order Value",   f"${avg_order:,.2f}",   "Calculated"],
        ["Unique Customers",      f"{unique_custs:,}",    "Active dataset"],
        ["Revenue per Customer",  f"${(total_rev / max(unique_custs,1)):,.2f}", "Calculated"],
        ["Orders per Customer",   f"{(total_orders / max(unique_custs,1)):.1f}", "Calculated"],
    ]
    if rating_col:
        avg_rating = float(df[rating_col].mean())
        kpi_data.append(["Avg Customer Rating", f"{avg_rating:.2f} / 5.0", "Calculated"])
    t_kpi = Table(kpi_data, colWidths=[230, 160, 150])
    t_kpi.setStyle(_ts(C_INDIGO))
    elems += [t_kpi, Spacer(1, 16)]

    # ── SECTION 2: AI INSIGHTS ───────────────────────────────────────────────
    elems.append(Paragraph("2. Strategic AI Insights", sec_style))
    ins_data = [["Priority", "Insight", "Recommended Action", "Confidence"]]
    for item in insights[:5]:
        p = item.get("priority", "–")
        col = {"High": C_INDIGO, "Medium": C_AMBER, "Low": C_GREEN}.get(p, C_TEXT)
        ins_data.append([
            Paragraph(f"<font color='#{_hex(col)}'><b>{p}</b></font>", body_style),
            Paragraph(item.get("title",    "–"), body_style),
            Paragraph(item.get("recommended_action", "–"), body_style),
            f"{item.get('confidence', 0):.0f}%",
        ])
    t_ins = Table(ins_data, colWidths=[58, 160, 250, 72])
    t_ins.setStyle(_ts(C_DARK))
    elems += [t_ins, Spacer(1, 16)]

    # ── SECTION 3: DASHBOARD CHARTS ──────────────────────────────────────────
    elems.append(PageBreak())
    elems.append(Paragraph("3. Dashboard Charts", sec_style))

    try:
        from services.analytics_service import (
            plot_revenue_trend, plot_category_revenue,
            plot_brand_performance, plot_region_sales,
            plot_top_products, plot_sales_funnel,
            plot_payment_methods, plot_age_distribution,
        )
        import concurrent.futures

        # Pre-calculate figures and submit to ThreadPool for parallel PNG generation
        figs = {
            "trend": (plot_revenue_trend(df), 900, 400),
            "cat": (plot_category_revenue(df), 440, 340),
            "brand": (plot_brand_performance(df), 440, 340),
            "reg": (plot_region_sales(df), 440, 340),
            "prod": (plot_top_products(df, top_n=8), 440, 340),
            "pay": (plot_payment_methods(df), 440, 340),
            "age": (plot_age_distribution(df), 440, 340),
            "funnel": (plot_sales_funnel(df), 700, 360),
        }

        pngs = {}
        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
            future_to_key = {
                executor.submit(_fig_to_png, fig, w, h): key
                for key, (fig, w, h) in figs.items()
            }
            for future in concurrent.futures.as_completed(future_to_key):
                key = future_to_key[future]
                try:
                    pngs[key] = future.result(timeout=60)
                except Exception as exc:
                    LOGGER.warning("Chart %s render failed: %s", key, exc)
                    pngs[key] = None

        # 3a. Revenue Trend (full width)
        elems.append(Paragraph("3a. Monthly Revenue Trend", sub_sec))
        png = pngs.get("trend")
        elems += _chart_elem(png, PAGE_W, 3.2 * inch)
        if png:
            elems.append(Paragraph(
                "Monthly revenue aggregated by calendar month from the active dataset.",
                caption_style,
            ))

        # 3b. Category Revenue + Brand Performance (side by side)
        elems.append(Paragraph("3b. Revenue by Category & Brand Performance", sub_sec))
        png_cat   = pngs.get("cat")
        png_brand = pngs.get("brand")
        half = (PAGE_W - 8) / 2
        row_data = [[]]
        if png_cat:
            row_data[0].append(Image(BytesIO(png_cat), width=half, height=2.7 * inch))
        else:
            row_data[0].append(Paragraph("Chart unavailable", body_style))
        if png_brand:
            row_data[0].append(Image(BytesIO(png_brand), width=half, height=2.7 * inch))
        else:
            row_data[0].append(Paragraph("Chart unavailable", body_style))
        side_tbl = Table(row_data, colWidths=[half, half])
        side_tbl.setStyle(TableStyle([
            ("LEFTPADDING",  (0,0),(-1,-1), 0),
            ("RIGHTPADDING", (0,0),(-1,-1), 0),
            ("TOPPADDING",   (0,0),(-1,-1), 0),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ]))
        elems.append(side_tbl)

        # 3c. Region Sales + Top Products (side by side)
        elems.append(Paragraph("3c. Regional Sales & Top Products", sub_sec))
        png_reg  = pngs.get("reg")
        png_prod = pngs.get("prod")
        row_data2 = [[]]
        if png_reg:
            row_data2[0].append(Image(BytesIO(png_reg), width=half, height=2.7 * inch))
        else:
            row_data2[0].append(Paragraph("Chart unavailable", body_style))
        if png_prod:
            row_data2[0].append(Image(BytesIO(png_prod), width=half, height=2.7 * inch))
        else:
            row_data2[0].append(Paragraph("Chart unavailable", body_style))
        side_tbl2 = Table(row_data2, colWidths=[half, half])
        side_tbl2.setStyle(TableStyle([
            ("LEFTPADDING",  (0,0),(-1,-1), 0),
            ("RIGHTPADDING", (0,0),(-1,-1), 0),
            ("TOPPADDING",   (0,0),(-1,-1), 0),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ]))
        elems.append(side_tbl2)

        # 3d. Payment Methods + Age Distribution (side by side)
        elems.append(PageBreak())
        elems.append(Paragraph("3d. Payment Methods & Customer Age Distribution", sub_sec))
        png_pay = pngs.get("pay")
        png_age = pngs.get("age")
        row_data3 = [[]]
        if png_pay:
            row_data3[0].append(Image(BytesIO(png_pay), width=half, height=2.7 * inch))
        else:
            row_data3[0].append(Paragraph("Chart unavailable", body_style))
        if png_age:
            row_data3[0].append(Image(BytesIO(png_age), width=half, height=2.7 * inch))
        else:
            row_data3[0].append(Paragraph("Chart unavailable", body_style))
        side_tbl3 = Table(row_data3, colWidths=[half, half])
        side_tbl3.setStyle(TableStyle([
            ("LEFTPADDING",  (0,0),(-1,-1), 0),
            ("RIGHTPADDING", (0,0),(-1,-1), 0),
            ("TOPPADDING",   (0,0),(-1,-1), 0),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ]))
        elems.append(side_tbl3)

        # 3e. Customer Conversion Funnel (full width)
        elems.append(Paragraph("3e. Customer Conversion Funnel", sub_sec))
        png_funnel = pngs.get("funnel")
        elems += _chart_elem(png_funnel, PAGE_W * 0.75, 2.8 * inch)

    except Exception as exc:
        LOGGER.warning("Charts section skipped: %s", exc)
        elems.append(Paragraph(
            f"Charts unavailable: {exc}. Ensure kaleido is installed (pip install kaleido).",
            body_style,
        ))

    # ── SECTION 4: MONTHLY REVENUE BREAKDOWN ─────────────────────────────────
    elems.append(PageBreak())
    elems.append(Paragraph("4. Monthly Revenue Breakdown", sec_style))
    if date_col and amt_col:
        tmp = df[[date_col, amt_col]].copy()
        tmp[date_col] = pd.to_datetime(tmp[date_col], errors="coerce")
        tmp[amt_col]  = pd.to_numeric(tmp[amt_col],   errors="coerce")
        tmp = tmp.dropna()
        if not tmp.empty:
            monthly = (
                tmp.groupby(pd.Grouper(key=date_col, freq="ME"))[amt_col]
                .agg(Revenue="sum", Orders="count")
                .reset_index()
            )
            monthly.columns = ["Month", "Revenue", "Orders"]
            monthly["Avg Order"] = (monthly["Revenue"] / monthly["Orders"].replace(0, 1)).round(2)
            monthly = monthly.sort_values("Month", ascending=False)
            monthly["Month"]   = monthly["Month"].dt.strftime("%b %Y")
            monthly["Revenue"] = monthly["Revenue"].apply(lambda v: f"${v:,.2f}")
            monthly["Avg Order"] = monthly["Avg Order"].apply(lambda v: f"${v:,.2f}")
            mon_data = [list(monthly.columns)] + monthly.values.tolist()
            t_mon = Table(mon_data, colWidths=[120, 160, 100, 160])
            t_mon.setStyle(_ts(C_GREEN))
            elems += [t_mon, Spacer(1, 16)]
    else:
        elems.append(Paragraph("Date/Amount columns not found for monthly breakdown.", body_style))

    # ── SECTION 5: CATEGORY SUMMARY TABLE ───────────────────────────────────
    if cat_col and amt_col:
        elems.append(Paragraph("5. Revenue by Category", sec_style))
        cat_df = df.groupby(cat_col)[amt_col].agg(["sum", "count", "mean"]).reset_index()
        cat_df.columns = ["Category", "Total Revenue", "Orders", "Avg Order Value"]
        cat_df = cat_df.sort_values("Total Revenue", ascending=False)
        cat_df["Total Revenue"]    = cat_df["Total Revenue"].apply(lambda v: f"${v:,.2f}")
        cat_df["Avg Order Value"]  = cat_df["Avg Order Value"].apply(lambda v: f"${v:,.2f}")
        cat_data = [list(cat_df.columns)] + cat_df.values.tolist()
        t_cat = Table(cat_data, colWidths=[160, 140, 100, 140])
        t_cat.setStyle(_ts(C_VIOLET))
        elems += [t_cat, Spacer(1, 16)]

    # ── SECTION 6: TOP 10 CUSTOMERS ──────────────────────────────────────────
    if cust_col and amt_col:
        elems.append(Paragraph("6. Top 10 Customers by Revenue", sec_style))
        top_custs = (
            df.groupby(cust_col)[amt_col]
            .sum()
            .sort_values(ascending=False)
            .head(10)
            .reset_index()
        )
        top_custs.columns = ["Customer ID", "Total Revenue"]
        top_custs.insert(0, "Rank", range(1, len(top_custs) + 1))
        top_custs["Total Revenue"] = top_custs["Total Revenue"].apply(lambda v: f"${v:,.2f}")
        tc_data = [list(top_custs.columns)] + top_custs.values.tolist()
        t_tc = Table(tc_data, colWidths=[60, 220, 180])
        t_tc.setStyle(_ts(C_AMBER))
        elems += [t_tc, Spacer(1, 16)]

    # ── SECTION 7: BRAND SUMMARY ─────────────────────────────────────────────
    if brand_col and amt_col:
        elems.append(Paragraph("7. Brand Performance Summary", sec_style))
        brand_df = df.groupby(brand_col)[amt_col].agg(["sum", "count"]).reset_index()
        brand_df.columns = ["Brand", "Total Revenue", "Orders"]
        brand_df = brand_df.sort_values("Total Revenue", ascending=False).head(15)
        brand_df["Total Revenue"] = brand_df["Total Revenue"].apply(lambda v: f"${v:,.2f}")
        br_data = [list(brand_df.columns)] + brand_df.values.tolist()
        t_br = Table(br_data, colWidths=[200, 160, 100])
        t_br.setStyle(_ts(colors.HexColor("#0EA5E9")))
        elems += [t_br, Spacer(1, 16)]

    # ── SECTION 8: TRANSACTION PREVIEW ───────────────────────────────────────
    elems.append(PageBreak())
    elems.append(Paragraph("8. Transaction Data Preview (Top 20 rows)", sec_style))
    show_cols = [c for c in
                 ["TransactionID", "CustomerID", "PurchaseDate", "Category",
                  "ProductName", "TotalAmount", "PaymentMethod"]
                 if c in df.columns]
    sample   = df[show_cols].head(20).copy()
    if date_col and date_col in sample.columns:
        sample[date_col] = pd.to_datetime(sample[date_col], errors="coerce").dt.strftime("%Y-%m-%d")
    if amt_col and amt_col in sample.columns:
        sample[amt_col] = sample[amt_col].apply(lambda v: f"${float(v):,.2f}")
    col_w = [int(PAGE_W / len(show_cols))] * len(show_cols)
    tbl_data = [list(sample.columns)] + [
        [str(v)[:22] for v in row] for _, row in sample.iterrows()
    ]
    t_txn = Table(tbl_data, colWidths=col_w)
    t_txn.setStyle(_ts(colors.HexColor("#334155")))
    elems.append(t_txn)

    # ── Footer note ──────────────────────────────────────────────────────────
    elems.append(Spacer(1, 20))
    elems.append(HRFlowable(width="100%", thickness=0.8, color=C_BORDER, spaceAfter=8))
    elems.append(Paragraph(
        f"<i>This report was auto-generated by {APP_NAME} on "
        f"{today.strftime('%B %d, %Y')}. "
        f"Data reflects the currently active filtered dataset ({len(df):,} rows).</i>",
        ParagraphStyle("Footer", parent=styles["Normal"],
                       fontName="Helvetica-Oblique", fontSize=8,
                       textColor=C_MUTED, alignment=1),
    ))

    doc.build(elems)
    buffer.seek(0)
    return buffer.getvalue()


# ─── Colour hex helper ────────────────────────────────────────────────────────

def _hex(c: Any) -> str:
    """Return 6-char hex string from a ReportLab colour (no leading #)."""
    try:
        r, g, b = int(c.red * 255), int(c.green * 255), int(c.blue * 255)
        return f"{r:02X}{g:02X}{b:02X}"
    except Exception:
        return "334155"


# ─── Excel Report ─────────────────────────────────────────────────────────────

def generate_excel_report(df: pd.DataFrame, insights: list[dict]) -> bytes:
    """
    Generate a multi-sheet Excel workbook with:
      Sheet 1 – Transaction Data
      Sheet 2 – Monthly Revenue
      Sheet 3 – Category Summary
      Sheet 4 – Brand Summary
      Sheet 5 – Regional Summary
      Sheet 6 – Customer Summary (top 50)
      Sheet 7 – AI Insights
    """
    wb = openpyxl.Workbook()

    # Styles
    fill_hdr    = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    fill_hdr2   = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_hdr3   = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    fill_hdr4   = PatternFill(start_color="A855F7", end_color="A855F7", fill_type="solid")
    fill_hdr5   = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    fill_hdr6   = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    fill_alt    = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    font_hdr    = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body   = Font(name="Segoe UI", size=10, color="1E293B")
    font_title  = Font(name="Segoe UI", size=14, bold=True, color="6366F1")
    thin = Border(
        left  =Side(style="thin", color="E2E8F0"),
        right =Side(style="thin", color="E2E8F0"),
        top   =Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def _auto_width(ws) -> None:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col[0].column)
            ].width = min(max_len + 4, 42)

    def _write_sheet(ws, frame: pd.DataFrame, hdr_fill=None, title: str = "") -> None:
        if hdr_fill is None:
            hdr_fill = fill_hdr
        if title:
            ws.append([title])
            title_cell = ws.cell(row=1, column=1)
            title_cell.font = font_title
            title_cell.alignment = left
            ws.append([])  # blank row
            start_row = 3
        else:
            start_row = 1
        for r_idx, row in enumerate(
            dataframe_to_rows(frame, index=False, header=True), start_row
        ):
            ws.append(row)
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(row=r_idx if title else r_idx, column=c_idx)
                if r_idx == start_row:
                    cell.fill      = hdr_fill
                    cell.font      = font_hdr
                    cell.alignment = center
                else:
                    cell.font      = font_body
                    cell.border    = thin
                    cell.alignment = left
                    if (r_idx - start_row) % 2 == 1:
                        cell.fill = fill_alt
        _auto_width(ws)

    # ── Sheet 1: Transaction Data ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Transaction Data"
    ws1.sheet_view.showGridLines = False
    _write_sheet(ws1, df, fill_hdr, title=f"{APP_NAME} – Transaction Data")

    # ── Sheet 2: Monthly Revenue ─────────────────────────────────────────────
    date_col = next((c for c in ["PurchaseDate", "Date", "OrderDate"] if c in df.columns), None)
    amt_col  = next((c for c in ["TotalAmount", "Revenue", "Amount"]  if c in df.columns), None)
    if date_col and amt_col:
        ws2 = wb.create_sheet("Monthly Revenue")
        ws2.sheet_view.showGridLines = False
        tmp = df[[date_col, amt_col]].copy()
        tmp[date_col] = pd.to_datetime(tmp[date_col], errors="coerce")
        tmp[amt_col]  = pd.to_numeric(tmp[amt_col],   errors="coerce")
        tmp = tmp.dropna()
        if not tmp.empty:
            monthly = (
                tmp.groupby(pd.Grouper(key=date_col, freq="ME"))[amt_col]
                .agg(Revenue="sum", Orders="count")
                .reset_index()
            )
            monthly.columns = ["Month", "Revenue", "Orders"]
            monthly["Avg Order Value"] = (monthly["Revenue"] / monthly["Orders"].replace(0, 1)).round(2)
            monthly["Month"] = monthly["Month"].dt.strftime("%b %Y")
            monthly = monthly.sort_values("Month")
            _write_sheet(ws2, monthly, fill_hdr3, title="Monthly Revenue Breakdown")

    # ── Sheet 3: Category Summary ─────────────────────────────────────────────
    cat_col = next((c for c in ["Category"] if c in df.columns), None)
    if cat_col and amt_col:
        ws3 = wb.create_sheet("Category Summary")
        ws3.sheet_view.showGridLines = False
        cat = df.groupby(cat_col).agg(
            Orders  =(amt_col, "count"),
            Revenue =(amt_col, "sum"),
            AvgBasket=(amt_col, "mean"),
        ).reset_index()
        cat.columns = ["Category", "Orders", "Revenue", "Avg Basket"]
        cat = cat.sort_values("Revenue", ascending=False)
        _write_sheet(ws3, cat, fill_hdr4, title="Revenue by Category")

    # ── Sheet 4: Brand Summary ────────────────────────────────────────────────
    brand_col = next((c for c in ["Brand"] if c in df.columns), None)
    if brand_col and amt_col:
        ws4 = wb.create_sheet("Brand Summary")
        ws4.sheet_view.showGridLines = False
        brand_df = df.groupby(brand_col).agg(
            Orders  =(amt_col, "count"),
            Revenue =(amt_col, "sum"),
            AvgBasket=(amt_col, "mean"),
        ).reset_index()
        brand_df.columns = ["Brand", "Orders", "Revenue", "Avg Basket"]
        brand_df = brand_df.sort_values("Revenue", ascending=False)
        _write_sheet(ws4, brand_df, fill_hdr6, title="Brand Performance")

    # ── Sheet 5: Regional Summary ─────────────────────────────────────────────
    region_col = next((c for c in ["Region", "Location", "State"] if c in df.columns), None)
    if region_col and amt_col:
        ws5 = wb.create_sheet("Regional Summary")
        ws5.sheet_view.showGridLines = False
        reg_df = df.groupby(region_col).agg(
            Orders  =(amt_col, "count"),
            Revenue =(amt_col, "sum"),
        ).reset_index()
        reg_df.columns = ["Region", "Orders", "Revenue"]
        reg_df = reg_df.sort_values("Revenue", ascending=False)
        _write_sheet(ws5, reg_df, fill_hdr2, title="Sales by Region")

    # ── Sheet 6: Customer Summary ─────────────────────────────────────────────
    cust_col = next((c for c in ["CustomerID"] if c in df.columns), None)
    if cust_col and amt_col:
        ws6 = wb.create_sheet("Customer Summary")
        ws6.sheet_view.showGridLines = False
        cust_agg: dict = {
            "TotalSpend" : (amt_col, "sum"),
            "OrderCount" : (amt_col, "count"),
            "AvgOrder"   : (amt_col, "mean"),
        }
        rating_col = next((c for c in ["CustomerRating", "Rating"] if c in df.columns), None)
        if rating_col:
            cust_agg["AvgRating"] = (rating_col, "mean")
        cust_summary = df.groupby(cust_col).agg(**cust_agg).reset_index()
        cust_summary = cust_summary.sort_values("TotalSpend", ascending=False).head(50)
        _write_sheet(ws6, cust_summary, fill_hdr5, title="Top 50 Customers by Spend")

    # ── Sheet 7: AI Insights ─────────────────────────────────────────────────
    if insights:
        ws7 = wb.create_sheet("AI Insights")
        ws7.sheet_view.showGridLines = False
        ins_df = pd.DataFrame([{
            "Priority":    i.get("priority",    ""),
            "Insight":     i.get("title",       ""),
            "Description": i.get("description", ""),
            "Action":      i.get("recommended_action", ""),
            "Impact":      i.get("impact",      ""),
            "Confidence":  f"{i.get('confidence', 0):.0f}%",
        } for i in insights])
        _write_sheet(ws7, ins_df, fill_hdr2, title="AI Strategic Insights")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ─── CSV Export ───────────────────────────────────────────────────────────────

def generate_csv_export(df: pd.DataFrame) -> bytes:
    """Return the dataframe as UTF-8 CSV bytes."""
    return df.to_csv(index=False).encode("utf-8")
